"""Repositorio Excel — carga y escritura del archivo de muestras por tenant.

Cada empresa tiene su propio XLSX bajo ``data/tenants/<sanitized>/``.
Optimizaciones (Parquet snapshot, append incremental) iguales que antes.
"""
from __future__ import annotations

import logging
import threading
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook

from src.application import IEquipoRepository, IMuestraRepository
from src.domain import Equipo, EstadoEquipo, Muestra
from src.infrastructure.settings import VARIABLES_ANALITICAS

logger = logging.getLogger("oilmine.excel")


def _normalize_for_parquet(df: pd.DataFrame) -> pd.DataFrame:
    """Coacciona columnas `object` mixtas a string para que Parquet las acepte."""
    df2 = df.copy()
    for col in df2.columns:
        if df2[col].dtype == "object" and col not in {"Equipo", "Estado"}:
            try:
                df2[col] = df2[col].astype("string")
            except Exception:
                df2[col] = df2[col].map(lambda x: None if pd.isna(x) else str(x))
    return df2


def _write_parquet(parquet_path: Path, cache_dir: Path, df: pd.DataFrame) -> None:
    """Guarda `df` como Parquet."""
    try:
        cache_dir.mkdir(parents=True, exist_ok=True)
        _normalize_for_parquet(df).to_parquet(parquet_path, index=False)
        logger.info(f"Snapshot Parquet guardado: {parquet_path.name} ({len(df)} filas)")
    except Exception as e:
        logger.warning(f"No se pudo escribir snapshot Parquet: {e}")


def _parquet_is_fresh(excel_path: Path, parquet_path: Path) -> bool:
    if not parquet_path.exists() or not excel_path.exists():
        return False
    return parquet_path.stat().st_mtime >= excel_path.stat().st_mtime


# ======================================================================
# ExcelManager — una instancia por tenant (path explícito)
# ======================================================================
class ExcelManager:
    """Lee Excel/Parquet una vez en memoria, apendiza incremental — thread-safe."""

    def __init__(self, excel_path: Path, sheet_name: str) -> None:
        self._excel_path = Path(excel_path)
        self._sheet_name = sheet_name
        self._io_lock = threading.Lock()
        self._df: Optional[pd.DataFrame] = None
        self._header: Optional[List[str]] = None

    @property
    def excel_path(self) -> Path:
        return self._excel_path

    @property
    def parquet_dir(self) -> Path:
        return self._excel_path.parent / ".cache"

    @property
    def parquet_path(self) -> Path:
        return self.parquet_dir / "data_flota.parquet"

    def _load(self) -> None:
        df: Optional[pd.DataFrame] = None
        pq = self.parquet_path
        xlsx = self._excel_path

        if _parquet_is_fresh(xlsx, pq):
            try:
                df = pd.read_parquet(pq)
                logger.info(f"Datos cargados desde Parquet ({len(df)} filas) — skip XLSX")
            except Exception as e:
                logger.warning(f"Parquet corrupto, recurro al XLSX: {e}")
                df = None

        if df is None:
            df = pd.read_excel(xlsx, sheet_name=self._sheet_name)
            df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce")
            df["Hora_Producto"] = pd.to_numeric(df["Hora_Producto"], errors="coerce")
            for col in VARIABLES_ANALITICAS:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors="coerce")
            _write_parquet(pq, self.parquet_dir, df)

        df = df.sort_values(["Equipo", "Fecha", "Hora_Producto"]).reset_index(drop=True)
        self._df = df
        self._header = list(df.columns)

    def _ensure_loaded(self) -> None:
        if self._df is None:
            self._load()

    def dataframe(self) -> pd.DataFrame:
        with self._io_lock:
            self._ensure_loaded()
            return self._df  # type: ignore[return-value]

    def preload(self) -> None:
        with self._io_lock:
            self._ensure_loaded()

    def rebuild_parquet(self) -> None:
        with self._io_lock:
            self._df = None
            self._load()

    def append_row(self, row: Dict[str, object]) -> None:
        """Añade una fila al Excel y al DF en memoria sin invalidar el cache."""
        with self._io_lock:
            wb = load_workbook(self._excel_path)
            ws = wb[self._sheet_name]
            header = [c.value for c in ws[1]]
            new_row = [row.get(col, None) for col in header]
            ws.append(new_row)
            wb.save(self._excel_path)
            wb.close()

            if self._df is not None:
                new_df_row = {col: row.get(col, None) for col in self._df.columns}
                if "Fecha" in new_df_row and new_df_row["Fecha"] is not None:
                    new_df_row["Fecha"] = pd.to_datetime(new_df_row["Fecha"], errors="coerce")
                if "Hora_Producto" in new_df_row:
                    new_df_row["Hora_Producto"] = pd.to_numeric(
                        new_df_row["Hora_Producto"], errors="coerce"
                    )
                for col in VARIABLES_ANALITICAS:
                    if col in new_df_row:
                        new_df_row[col] = pd.to_numeric(new_df_row[col], errors="coerce")
                self._df = pd.concat(
                    [self._df, pd.DataFrame([new_df_row])],
                    ignore_index=True,
                )
                self._df = self._df.sort_values(
                    ["Equipo", "Fecha", "Hora_Producto"]
                ).reset_index(drop=True)

                try:
                    _write_parquet(self.parquet_path, self.parquet_dir, self._df)
                except Exception:
                    pass
            else:
                self._df = None


# ======================================================================
# Repositorios
# ======================================================================
class ExcelEquipoRepository(IEquipoRepository):
    def __init__(self, manager: ExcelManager) -> None:
        self.manager = manager

    def listar_ids(self) -> List[str]:
        df = self.manager.dataframe()
        return sorted(df["Equipo"].dropna().unique().tolist())

    def obtener(self, equipo_id: str) -> Equipo:
        df = self.manager.dataframe()
        grp = (
            df[df["Equipo"] == equipo_id]
            .sort_values(["Fecha", "Hora_Producto"])
            .reset_index(drop=True)
        )
        if grp.empty:
            raise ValueError(f"Equipo '{equipo_id}' no existe en el Excel")

        return self._build_equipo(equipo_id, grp)

    def obtener_todos(self) -> List[Equipo]:
        df = self.manager.dataframe()
        equipos: List[Equipo] = []
        for equipo_id, grp in df.groupby("Equipo"):
            grp = grp.sort_values(["Fecha", "Hora_Producto"]).reset_index(drop=True)
            equipos.append(self._build_equipo(str(equipo_id), grp))
        return equipos

    def _build_equipo(self, equipo_id: str, grp: pd.DataFrame) -> Equipo:
        muestras: List[Muestra] = []
        for _, r in grp.iterrows():
            estado_raw = r.get("Estado")
            try:
                estado = EstadoEquipo(estado_raw) if pd.notna(estado_raw) else None
            except ValueError:
                estado = None

            variables: Dict[str, float] = {}
            for v in VARIABLES_ANALITICAS:
                val = r.get(v)
                if pd.notna(val):
                    variables[v] = float(val)
                else:
                    variables[v] = float("nan")

            hora = r.get("Hora_Producto")
            if pd.isna(hora):
                continue
            fecha_val = r.get("Fecha")
            fecha_obj = (
                fecha_val.date() if isinstance(fecha_val, (pd.Timestamp, datetime)) else None
            )

            muestras.append(
                Muestra(
                    equipo=equipo_id,
                    fecha=fecha_obj,
                    hora_producto=float(hora),
                    estado=estado,
                    variables=variables,
                )
            )
        return Equipo(id=equipo_id, muestras=muestras)


class ExcelMuestraRepository(IMuestraRepository):
    def __init__(self, manager: ExcelManager) -> None:
        self.manager = manager

    def registrar(self, muestra: Muestra) -> None:
        fecha = muestra.fecha
        if isinstance(fecha, date) and not isinstance(fecha, datetime):
            fecha = datetime.combine(fecha, datetime.min.time())

        row = {
            "Equipo": muestra.equipo,
            "Fecha": fecha,
            "Hora_Producto": muestra.hora_producto,
        }
        if muestra.estado is not None:
            row["Estado"] = muestra.estado.value
        row.update(muestra.variables)
        self.manager.append_row(row)
